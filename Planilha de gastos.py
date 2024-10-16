import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import tkinter as tk
from tkinter import messagebox

# Inicializa uma lista para armazenar os gastos
gastos = []


def adicionar_gasto():
    """Adiciona um gasto à lista de gastos a partir dos campos de entrada."""
    data = entry_data.get()
    descricao = entry_descricao.get()
    valor = entry_valor.get()

    try:
        valor = float(valor)
        gastos.append({"Data": data, "Descrição": descricao, "Valor": valor})
        messagebox.showinfo("Sucesso", "Gasto adicionado!")
        entry_data.delete(0, tk.END)
        entry_descricao.delete(0, tk.END)
        entry_valor.delete(0, tk.END)
    except ValueError:
        messagebox.showerror("Erro", "Por favor, insira um valor numérico válido.")


def salvar_planilha():
    """Salva a lista de gastos em uma planilha Excel com estilo."""
    if not gastos:
        messagebox.showwarning("Aviso", "Nenhum gasto registrado!")
        return

    df = pd.DataFrame(gastos)

    # Criar um Workbook e adicionar uma planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    # Definir a cor cinza
    cinza_claro = "D9D9D9"

    # Adicionar cabeçalhos com cor cinza
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title).fill = PatternFill(start_color=cinza_claro,
                                                                              end_color=cinza_claro, fill_type="solid")

    # Adicionar dados e bordas
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border

    # Calcular a somatória dos gastos
    total_gastos = df["Valor"].sum()
    ws.cell(row=len(gastos) + 2, column=2, value="Total:").alignment = Alignment(horizontal='right')
    ws.cell(row=len(gastos) + 2, column=3, value=total_gastos).border = border

    # Salvar o arquivo
    nome_arquivo = "gastos.xlsx"
    wb.save(nome_arquivo)
    messagebox.showinfo("Sucesso", f"Planilha salva como '{nome_arquivo}'.")


# Configuração da interface gráfica
root = tk.Tk()
root.title("Controle de Gastos")

# Criação dos campos de entrada
tk.Label(root, text="Data (YYYY-MM-DD):").grid(row=0, column=0)
entry_data = tk.Entry(root)
entry_data.grid(row=0, column=1)

tk.Label(root, text="Descrição:").grid(row=1, column=0)
entry_descricao = tk.Entry(root)
entry_descricao.grid(row=1, column=1)

tk.Label(root, text="Valor:").grid(row=2, column=0)
entry_valor = tk.Entry(root)
entry_valor.grid(row=2, column=1)

# Criação dos botões
tk.Button(root, text="Adicionar Gasto", command=adicionar_gasto).grid(row=3, column=0, columnspan=2)
tk.Button(root, text="Salvar Planilha", command=salvar_planilha).grid(row=4, column=0, columnspan=2)

# Iniciar a interface
root.mainloop()
